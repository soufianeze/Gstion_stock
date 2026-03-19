"""
CVE — Serveur local de gestion de stock
Lit et écrit directement dans CVE_Stock.xlsx
Lancer : python serveur.py
Ouvrir  : http://localhost:5000
"""
import os, json
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='.')

# ── Chemin du fichier Excel (même dossier que serveur.py)
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'CVE_Stock.xlsx')

# ── Helpers Excel
def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="1C2B3A", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def now_str():
    return datetime.now().strftime("%d/%m/%Y %H:%M")

# ══════════════════════════════════════════════════════
# LIRE depuis Excel
# ══════════════════════════════════════════════════════
def init_excel():
    """Crée un fichier Excel vide s'il n'existe pas."""
    if not os.path.exists(EXCEL_FILE):
        ecrire_excel([], [])

def lire_excel():
    init_excel()
    if not os.path.exists(EXCEL_FILE):
        return [], []

    wb = load_workbook(EXCEL_FILE, data_only=True)

    # ── Produits
    produits = []
    if "Stock" in wb.sheetnames:
        ws = wb["Stock"]
        # Lire en-têtes ligne 1
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column - 1
        # Lire données ligne 2+
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[headers.get("Désignation", 1)]:
                continue
            def g(key, default=""):
                idx = headers.get(key)
                return row[idx] if idx is not None and idx < len(row) else default
            produits.append({
                "ref":       str(g("Référence", "") or ""),
                "nom":       str(g("Désignation", "") or ""),
                "cat":       str(g("Catégorie", "Autre") or "Autre"),
                "unite":     str(g("Unité", "Pièce") or "Pièce"),
                "qte":       float(g("Quantité", 0) or 0),
                "min":       float(g("Min", 0) or 0),
                "max":       float(g("Max", 0) or 0),
                "prix":      float(g("Prix", 0) or 0),
                "desc":      str(g("Description", "") or ""),
                "createdAt": str(g("Date Création", now_str()) or now_str()),
                "updatedAt": str(g("Dernière MAJ", now_str()) or now_str()),
            })

    # ── Mouvements
    mouvements = []
    if "Mouvements" in wb.sheetnames:
        ws2 = wb["Mouvements"]
        headers2 = {}
        for cell in ws2[1]:
            if cell.value:
                headers2[str(cell.value).strip()] = cell.column - 1
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row: continue
            def g2(key, default=""):
                idx = headers2.get(key)
                return row[idx] if idx is not None and idx < len(row) else default
            if not g2("Type"): continue
            mouvements.append({
                "date":       str(g2("Date & Heure", now_str())),
                "type":       "entree" if "ntr" in str(g2("Type","")).lower() else "sortie",
                "produit":    str(g2("Désignation", "")),
                "qte":        float(g2("Quantité", 0) or 0),
                "stockAvant": float(g2("Stock Avant", 0) or 0),
                "stockApres": float(g2("Stock Après", 0) or 0),
                "unite":      str(g2("Unité", "")),
                "resp":       str(g2("Responsable", "")),
                "note":       str(g2("Note", "")),
            })

    wb.close()
    return produits, mouvements


# ══════════════════════════════════════════════════════
# ÉCRIRE dans Excel
# ══════════════════════════════════════════════════════
def ecrire_excel(produits, mouvements):
    wb = Workbook()

    # ── Feuille Stock
    ws = wb.active
    ws.title = "Stock"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    HDRS = [
        ("Référence", 11), ("Désignation", 28), ("Catégorie", 15),
        ("Unité", 9), ("Quantité", 12), ("Min", 9), ("Max", 9),
        ("Prix", 11), ("Valeur (DH)", 13), ("Statut", 13),
        ("Date Création", 20), ("Dernière MAJ", 20), ("Description", 25),
    ]
    for i, (label, width) in enumerate(HDRS, 1):
        c = ws.cell(1, i, label)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = fill("1A4A6B")
        c.alignment = aln()
        c.border = bdr()
        ws.column_dimensions[c.column_letter].width = width
    ws.row_dimensions[1].height = 24

    for i, p in enumerate(produits):
        r = i + 2
        statut = "Critique" if float(p.get("qte",0)) <= 0 or float(p.get("qte",0)) < float(p.get("min",0))*.5 \
                 else "Faible" if float(p.get("qte",0)) < float(p.get("min",0)) else "Normal"
        valeur = float(p.get("qte",0)) * float(p.get("prix",0))
        vals = [
            p.get("ref",""), p.get("nom",""), p.get("cat",""), p.get("unite",""),
            float(p.get("qte",0)), float(p.get("min",0)), float(p.get("max",0)),
            float(p.get("prix",0)), valeur, statut,
            p.get("createdAt", now_str()), p.get("updatedAt", now_str()), p.get("desc",""),
        ]
        bg = "E6F5F4" if i % 2 == 0 else "FDFAF5"
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.fill = fill(bg); cell.border = bdr()
            cell.font = fnt(size=10); cell.alignment = aln("left")
        ws.cell(r, 5).font = fnt(bold=True, size=12); ws.cell(r, 5).alignment = aln()
        ws.cell(r, 8).number_format = '#,##0.00 "DH"'
        ws.cell(r, 9).number_format = '#,##0.00 "DH"'
        # Colorer statut
        sc = {"Normal":"2E8B57","Faible":"D4860A","Critique":"E05A3A"}.get(statut,"1C2B3A")
        ws.cell(r, 10).font = fnt(bold=True, color=sc); ws.cell(r, 10).alignment = aln()
        ws.cell(r, 11).font = fnt(italic=True, color="8FA8BC", size=9)
        ws.cell(r, 12).font = fnt(italic=True, color="8FA8BC", size=9)
        ws.row_dimensions[r].height = 18

    # Ligne totaux
    tr = len(produits) + 2
    ws.merge_cells(f"A{tr}:D{tr}")
    ws.cell(tr, 1, "TOTAUX").font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws.cell(tr, 1).fill = fill("1A4A6B"); ws.cell(tr, 1).alignment = aln()
    total_qte = sum(float(p.get("qte",0)) for p in produits)
    total_val = sum(float(p.get("qte",0))*float(p.get("prix",0)) for p in produits)
    ws.cell(tr, 5, total_qte).font = fnt(bold=True, color="1A4A6B")
    ws.cell(tr, 9, total_val).font = fnt(bold=True, color="1A4A6B")
    ws.cell(tr, 9).number_format = '#,##0.00 "DH"'
    for col in range(1, 14):
        ws.cell(tr, col).fill = fill("E8F0F7")
    ws.row_dimensions[tr].height = 20

    # ── Feuille Mouvements
    ws2 = wb.create_sheet("Mouvements")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    MVT_HDRS = [
        ("Date & Heure",20),("Jour",12),("Type",11),("Désignation",26),
        ("Quantité",11),("Unité",9),("Stock Avant",12),("Stock Après",12),
        ("Responsable",16),("Note",28),
    ]
    for i,(label,width) in enumerate(MVT_HDRS,1):
        c = ws2.cell(1,i,label)
        c.font = Font(name="Arial",bold=True,size=10,color="FFFFFF")
        c.fill = fill("3AAFA9"); c.alignment = aln(); c.border = bdr()
        ws2.column_dimensions[c.column_letter].width = width
    ws2.row_dimensions[1].height = 24

    jours = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
    for i, m in enumerate(mouvements):
        r = i + 2
        try:
            # Tenter de parser la date pour obtenir le jour
            for fmt in ["%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d"]:
                try: jour = jours[datetime.strptime(str(m.get("date",""))[:19].replace("T"," "), fmt).weekday()]; break
                except: jour = ""
        except: jour = ""
        typ_label = "Entrée" if m.get("type") == "entree" else "Sortie"
        vals = [m.get("date",""), jour, typ_label, m.get("produit",""),
                m.get("qte",0), m.get("unite",""), m.get("stockAvant",0),
                m.get("stockApres",0), m.get("resp",""), m.get("note","")]
        bg = "E6F5F4" if i%2==0 else "FDFAF5"
        for j,v in enumerate(vals,1):
            cell=ws2.cell(r,j,v); cell.fill=fill(bg); cell.border=bdr()
            cell.font=fnt(size=10); cell.alignment=aln("left")
        tc = "2E8B57" if m.get("type")=="entree" else "E05A3A"
        ws2.cell(r,3).font = Font(name="Arial",bold=True,color=tc,size=10)
        ws2.row_dimensions[r].height = 18

    # ── Feuille Journal
    ws3 = wb.create_sheet("Journal")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"
    JRN_HDRS = [("Date & Heure",20),("Jour",12),("Type",17),("Objet",26),("Détail",35),("Opérateur",16)]
    for i,(label,width) in enumerate(JRN_HDRS,1):
        c=ws3.cell(1,i,label)
        c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
        c.fill=fill("E8A838"); c.alignment=aln(); c.border=bdr()
        ws3.column_dimensions[c.column_letter].width=width
    ws3.row_dimensions[1].height=24

    wb.save(EXCEL_FILE)
    wb.close()


# ══════════════════════════════════════════════════════
# API REST
# ══════════════════════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory('.', 'interface.html')

@app.route('/api/produits', methods=['GET'])
def get_produits():
    produits, mouvements = lire_excel()
    return jsonify({"produits": produits, "mouvements": mouvements, "fichier": EXCEL_FILE})

@app.route('/api/produits', methods=['POST'])
def save_produits():
    data = request.json
    produits   = data.get("produits", [])
    mouvements = data.get("mouvements", [])
    try:
        ecrire_excel(produits, mouvements)
        return jsonify({"ok": True, "message": f"✅ {len(produits)} produits sauvegardés dans Excel", "fichier": EXCEL_FILE, "date": now_str()})
    except Exception as e:
        return jsonify({"ok": False, "message": f"❌ Erreur : {str(e)}"}), 500

@app.route('/api/statut', methods=['GET'])
def statut():
    existe = os.path.exists(EXCEL_FILE)
    taille = os.path.getsize(EXCEL_FILE) if existe else 0
    modif  = datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).strftime("%d/%m/%Y %H:%M") if existe else "—"
    return jsonify({"fichier": EXCEL_FILE, "existe": existe, "taille_ko": round(taille/1024,1), "modifie": modif})


# ══════════════════════════════════════════════════════
# LANCEMENT
# ══════════════════════════════════════════════════════
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
